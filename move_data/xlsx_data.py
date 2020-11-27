from openpyxl import load_workbook


class ReadXLSX:
    def __init__(self, xlsx_filepath, sheet):
        """Читаем xlsx-файл построчно."""
        def iter_rows(worksheet):
            for row in ws.iter_rows():
                yield [cell.value for cell in row]

        wb = load_workbook(xlsx_filepath)
        ws = wb[sheet]

        row_data = list(iter_rows(worksheet=ws))

        # Данные читаются с первой строки.
        self.xlsx_data = row_data

        # Данные читаются со второй строки.
        # self.xlsx_data = row_data[1:]
